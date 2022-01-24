# Data Import - Implementation
# ----------------------------------------------------

import glob
import os
import openpyxl
import requests
import json
from datetime import date
import configparser

# Read configurations
config = configparser.ConfigParser()		
config.read("config.ini")


# Initialization
server = config['SERVER']
host = "{}://{}:{}".format(server['protocol'], server['host'], server['port'])
f = open(".token","r")
endpoint = "{}/api/v1/transactions".format(host)
headers = {
    "Authorization": "Bearer {}".format("".join(f.readlines())),
    "Content-Type": "application/json",
    "Accept": "application/json"
}


# Get XLSX files inside of data directory
print("\nChecking xlsx files inside of data directory...")
os.chdir("data")
files = []
for file in glob.glob("*.xlsx"):
    if file.startswith('~$'): continue
    files.append(file)
print("> Found {} file(s).\n".format(len(files)))


# Extract transactions from files
print("Retrieving transactions...")
transactionList = []
for file in files:
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    rowIndex = 7
    isTransaction = sheet.cell(row = rowIndex, column = 1).value != None
    while isTransaction:

        amount = sheet.cell(row = rowIndex, column = 4).value
        transactionType = 0
        if amount == None:
            transactionType = 1
            amount = sheet.cell(row = rowIndex, column = 5).value

        transaction = {
            'date': sheet.cell(row = rowIndex, column = 1).value,
            'description': sheet.cell(row = rowIndex, column = 3).value.strip(),
            'type': transactionType,
            'amount': amount
        }
        
        if transaction not in transactionList: 
            transactionList.append(transaction)
        
        rowIndex += 1
        isTransaction = sheet.cell(row = rowIndex, column = 1).value != None

print("> Captured {} transactions from file.\n".format(len(transactionList)))



# Analysing transactions, filtering by transactions already registered
print("Analysing transactions...")
transactionsNotCreated = []
for transaction in transactionList:

    exist = False
    params = {
        'start': transaction["date"].date(),
        'end': transaction["date"].date()
    }
    data = requests.get(endpoint, params=params, headers=headers).json()
    
    for d in data["data"]:
        info = d["attributes"]["transactions"][0]
        if info["description"].strip() == transaction["description"] and float(info["amount"]) == transaction["amount"]:
            exist = True
            break

    if exist is False:            
        transactionsNotCreated.append(transaction)
        
if len(transactionsNotCreated) == 0:
    print("> Transactions already created. Nothing to do.\n")
    exit(0)
elif len(transactionsNotCreated) == len(transactionList):
    print("> Detected {} transactions to register.\n".format(len(transactionsNotCreated)))
else:
    print("> Detected {} transactions to register ({} transactions are already registered).\n".format(len(transactionsNotCreated), len(transactionList) - len(transactionsNotCreated)))


# Check user acceptance
accept=input("Do you want to proceed with transation registration? [y/n]: ") 
if accept.lower() != 'y':
    print("> Transaction registration cancelled.\n") 
    exit(0)


# Prepare to create transactions 
print("Preparing to register transactions...")
for transaction in transactionsNotCreated:

    print("> [{}] {}: {}".format(transaction["date"].date(), transaction["description"], transaction['amount']), end=" >> ")
    data = {
        'account_id': 1,
        'date': transaction["date"].isoformat(),
        'description': transaction["description"],
        'type': 'withdrawal' if transaction["type"] == 0 else 'deposit',
        'amount': transaction["amount"],
        'source_id': 1 if transaction["type"] == 0 else None,
        'destination_id': 1 if transaction["type"] == 1 else None
    }

    data = {
        'error_if_duplicate_hash': False,
        'apply_rules': True,
        'transactions': [ data ] 
    }

    r = requests.post(endpoint, data=json.dumps(data), headers=headers)
    print("200 OK") if r.status_code == 200 else print(r.status_code + " Error")

        
    
print("\n")